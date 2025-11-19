from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Sum
from django.conf import settings

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from account.models import User,Student,Administration
from course.models import Course, Course_registration, Course_project
from repo.models import Repo_commit, Repo_pr ,Repo_issue, Repository,Repo_contributor
from itertools import groupby
from operator import itemgetter
from collections import defaultdict
import statistics
import numpy as np
from openpyxl.styles import Alignment
from io import BytesIO
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
import io

import json
from django.db.models import Prefetch, Count, Q, F, Min, Max
from django.http import JsonResponse

import requests
from account.api.views import get_kuopenapi_access_token,query_student_openapi,student_read_course_info

class HealthCheckAPIView(APIView):
    def get(self, request):
        return Response({"status": "OK"}, status=status.HTTP_200_OK)

# ========================================
# Frontend Function
# ========================================
def course_create_db (request):
    try:
        course_id=request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')
        name = request.GET.get('name')
        prof = request.GET.get('prof')
        ta = request.GET.get('ta')
        student_count = request.GET.get('student_count')

        if not course_id or not year or not semester:
            raise ValueError("course_id and year and semester cannot be empty")
        
        if Course.objects.filter(course_id=course_id, year=year,semester=semester).exists():
            raise ValueError("course already exists.")
        
        course = Course.objects.create(
            course_id=course_id,
            year=year,
            semester=semester,
            name=name,
            prof=prof,
            ta=ta,
            student_count=student_count
        )

        return JsonResponse({"status": "OK", "message": "course record created successfully"})
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)     

def course_read_db(request):
    try:
        # 1) 필요한 모든 데이터를 한 번에 가져오기 (select_related, prefetch_related)
        courses = Course.objects.select_related().prefetch_related(
            Prefetch(
                'course_project_set',
                queryset=Course_project.objects.select_related('repo').filter(
                    repo__forked=False
                )
            ),
            Prefetch(
                'course_registration_set',
                queryset=Course_registration.objects.select_related('student')
            )
        ).all()
        
        data = []
        
        for course in courses:
            # 2) 이미 prefetch된 데이터 사용 (추가 쿼리 없음)
            course_projects = course.course_project_set.all()
            
            # 3) DB aggregation 활용
            stats = course_projects.aggregate(
                total_commits=Sum('repo__commit_count'),
                total_issues=Sum(
                    F('repo__open_issue_count') + F('repo__closed_issue_count')
                ),
                total_prs=Sum(
                    F('repo__open_pr_count') + F('repo__closed_pr_count')
                ),
                total_stars=Sum('repo__star_count'),
                repo_count=Count('id')
            )
            
            student_count = course.course_registration_set.count()
            
            # contributor 계산 (이 부분만 Python 처리)
            contributor_count = 0
            for cp in course_projects:
                if cp.repo.contributors:
                    contributor_count += cp.repo.contributors.count(',') + 1
            
            data.append({
                "course_id": course.course_id,
                "year": course.year,
                "semester": course.semester,
                "name": course.name,
                "prof": course.prof,
                "ta": course.ta,
                "student_count": student_count,
                "total_commits": stats['total_commits'] or 0,
                "total_issues": stats['total_issues'] or 0,
                "total_prs": stats['total_prs'] or 0,
                "total_stars": stats['total_stars'] or 0,
                "avg_commits": round((stats['total_commits'] or 0) / student_count, 2) if student_count > 0 else 0,
                "repository_count": stats['repo_count'] or 0,
                "contributor_count": contributor_count
            })
        
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_update_db(request):
    try:
        return JsonResponse({"status": "OK", "message": "course record update successfully"})
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)       

def course_delete_db(request):
    try:
        course_id=request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')           
        course = Course.objects.get(course_id=course_id, year=year, semester=semester)
        course.delete()  # 사용자 객체를 삭제합니다.
        return JsonResponse({"status": "OK", "message": "course record deleted successfully"})
    
    except ObjectDoesNotExist:
        return JsonResponse({"status": "Error", "message": f"Course '{course_id}' does not exist"}, status=404)    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)



def course_registration_create_db(request):
    
    try:
        course_id = request.GET.get('course_id')
        year = request.GET.get('year')
        semester = request.GET.get('semester')  
        student_ids = Student.objects.values_list('id', flat=True)
        course = Course.objects.get(course_id=course_id, year=year, semester=semester)
        
        count =0 

        for student_id in student_ids:
            student = Student.objects.get(id=student_id)    
            count+=1
            # 이미 해당 수업에 등록된 학생이 있는지 확인
            if Course_registration.objects.filter(course=course, student=student).exists():
                print(f"{student} already registered for {course.name}")
                continue
            
            course_reg = Course_registration.objects.create(
                course=course,  
                course_year=course.year,
                course_semester=course.semester,
                student=student
            )
            print(f"OK! {student} has been registered for {course.name}")
            
        print(f"The num of students:{count}")
        course.student_count = count
        course.save()
        return JsonResponse({"status": "OK", "message": "course_registration record created successfully"})
    
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

    
def course_project_update(request):
    try:

        courses = Course.objects.all()

        for course in courses:
            repos = Repository.objects.filter(name__icontains=course.course_repo_name)

            for repo in repos:
                # 이미 해당 course_id, year, semester, repo_id를 가지는 레코드가 있는지 확인
                if Course_project.objects.filter(course=course, course_year=course.year, course_semester=course.semester, repo_id=repo.id).exists():
                    print(f"{repo.owner_github_id}'s {repo.name} Course projects already exists!")
                    continue
                
                # 존재하지 않는다면 새로운 레코드 생성
                Course_project.objects.create(  
                    course=course,
                    course_year=course.year,
                    course_semester=course.semester,
                    repo_id=repo.id,
                    repo_name=repo.name
                )
                print(f"{repo.owner_github_id}'s {repo.name} has been created!")

        
        return JsonResponse({"status": "OK", "message": "course_project record created successfully"})
    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_year_search(request):
    try:
        data = []
        year = request.GET.get('year')
        students = Student.objects.all()
        
        for student in students:
            # 특정 학생의 모든 레포지토리 가져옴
            student_repos = Repository.objects.filter(owner_github_id=student.github_id)
            
            # 특정 학생의 과목 관련된 레포지토리들을 가져옴
            courses_repos = Course_project.objects.filter(repo__in=student_repos)
            
            # 특정 학생이 듣는 모든 course_id 테이블 및 딕셔너리 생성
            course_ids = []
            for course_project in courses_repos:
                course_id = course_project.course.course_id
                if course_id not in course_ids:
                    course_ids.append(course_id)

            # 딕셔너리 초기화
            course_dict = {course_id: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0} for course_id in course_ids}
            etc_dict = {year: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0}}

            courses_project_repos = [c.repo for c in courses_repos]
            
            print(courses_project_repos)

            # 각 repo에 대한 정보 추가
            for repo in student_repos:
                repo_created_year = str(datetime.strptime(repo.created_at, '%Y-%m-%dT%H:%M:%SZ').year)  # 연도 추출
                if repo_created_year == year:
                    
                    if repo in courses_project_repos:  # 특정 학생의 repo가 과목과 관련이 있는 경우
                        specific_course = Course_project.objects.get(repo=repo)
                        course_id = specific_course.course.course_id
                    
                        # Commit 수 합산
                        course_dict[course_id]['commit'] += repo.commit_count
                        
                        # PR 수 합산
                        course_dict[course_id]['pr'] += Repo_pr.objects.filter(repo_id=repo.id).count()
                        
                        # Issue 수 합산
                        course_dict[course_id]['issue'] += Repo_issue.objects.filter(repo_id=repo.id).count()
                        
                        # Repo 수 합산
                        course_dict[course_id]['repo'] += 1
                    else:  # 특정 학생의 repo가 과목과 관련 없는 경우
                        etc_dict[repo_created_year]['commit'] += repo.commit_count
                        etc_dict[repo_created_year]['pr'] += Repo_pr.objects.filter(repo_id=repo.id).count()
                        etc_dict[repo_created_year]['issue'] += Repo_issue.objects.filter(repo_id=repo.id).count()
                        etc_dict[repo_created_year]['repo'] += 1

                else: 
                    continue 
            # 각 과목에 대한 정보 추가
            
        for course_id, course_count in course_dict.items():
                course = Course.objects.get(course_id=course_id)
                if str(course.year) == year :

                    course_info = {
                        "year": course.year,
                        "semester": course.semester,
                        "course_name": course.name,
                        "commit": course_count['commit'],
                        "pr": course_count['pr'],
                        "issue": course_count['issue'],
                        "num_repos": course_count['repo']
                    }

                    data.append(course_info)
        
        for the_year, the_year_count in etc_dict.items():
            etc_info = {
                "year": the_year,
                "semester": "",
                "course_name": the_year + " 기타",
                "commit": the_year_count['commit'],
                "pr": the_year_count['pr'],
                "issue": the_year_count['issue'],
                "num_repos": the_year_count['repo']
            }

            data.append(etc_info)               
                        
        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)
    

def course_department_count(request):
    try:
        data = []
        # 해당 년도의 모든 과목 코드들 가져옴
        
        # 간단하게 생각스.. 
        distinct_course_ids = Course.objects.all()
        print(distinct_course_ids)
        # Initialize the dictionary to store department counts

        for specific_course_id in distinct_course_ids:
            
            department_count = {}
            
            specific_course_reg = Course_registration.objects.filter(course = specific_course_id)

            for specific_course_per_reg in specific_course_reg : 
                student = Student.objects.get(id = specific_course_per_reg.student.id)
                
                # Get the department of the student
                department_name = student.department
                
                # Check if the department already exists in the dictionary
                if department_name in department_count:
                    # If it exists, increment the count
                    department_count[department_name] += 1
                else:
                    # If it doesn't exist, add it with a count of 1
                    department_count[department_name] = 1
            
            # You can now return this dictionary as part of the response, or process it further
            data.append({'year': specific_course_id.year,'semester': specific_course_id.semester ,'course_id': specific_course_id.course_id,'course_name': specific_course_id.name, 'department_count': department_count , "total_count": specific_course_id.student_count})
        
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

                


def course_validation(request):
    try:
        courses = Course.objects.all()

        for specific_course in courses:
            # 과목 학생 수 count validation 
            specific_course_count = Course_registration.objects.filter(course=specific_course).count()
            specific_course.student_count = specific_course_count
            print(f'{specific_course.name} {specific_course.course_id} student count has been changed to {specific_course.student_count}')
            specific_course.save()

        return JsonResponse({'status': 'success', 'message': 'Course student count updated'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def course_reg_validation(request):
    try:
        access_token = get_kuopenapi_access_token()
        courses = Course.objects.all()
        data = []
        for specific_course in courses:
            course_id, course_class = specific_course.course_id.split('-')
            
            #API 요청
            api_url = "https://kuapi.korea.ac.kr/svc/course/course-registration/by-class"  # 실제 API 엔드포인트로 변경
            headers = {
                'AUTH_KEY': access_token
            }
            
            # 요청 파라미터 설정
            params = {
                'client_id': settings.KOREAUNIV_OPENAPI_CLIENT_ID,
                "year":str(specific_course.year),
                'term':str(specific_course.semester)+'R',
                'course_code': course_id,
                'course_class': course_class
            }
            
            # API 호출
            response = requests.get(api_url, headers=headers, params=params)
            # JSON 응답을 파싱
            
            response_data = response.json()
            
            result_items = response_data.get("result", [])
            
            new_ids = []

            for result_item in result_items:

                student_reg_id = result_item.get("student_id")
                
                try:
                    # Student 객체를 조회
                    reg_student = Student.objects.get(id=student_reg_id)
                    
                    try:
                        # Course_registration 객체가 존재하는지 확인
                        Course_registration.objects.get(student=reg_student, course = specific_course)
                        # 존재하면 pass
                        pass
                    
                    except Course_registration.DoesNotExist:
                        # Course_registration 객체가 존재하지 않으면 student_reg_id를 추가
                        Course_registration.objects.create(course = specific_course, course_year = specific_course.year , course_semester = specific_course.semester , student = reg_student)
                
                except Student.DoesNotExist:
                    # Student 객체가 존재하지 않으면 student_reg_id를 추가
                    new_ids.append(student_reg_id)

            
            #새로운 (미등록)학생들에 대해서 학생DB에 추가
            new_student_data,empty_ids = query_student_openapi(new_ids,access_token)
            print([specific_course.name,new_student_data,"empty",empty_ids])
            new_count = 0 

            for item in new_student_data :
                try :
                    id = item.get("STD_ID")
                    enrollment = item.get("REC_STS_NM")
                    name = item.get("KOR_NM")
                    college = item.get("COL_NM")
                    department = item.get("DEPT_NM")
                    double_major = item.get("SMAJOR_NM")
                    email_addr = item.get("email_addr")

                    new_reg_student, created = Student.objects.get_or_create(
                        id=id,                   
                        enrollment =enrollment,
                        name = name,
                        college = college,
                        department = department,
                        double_major = double_major,
                        primary_email = email_addr
                        
                    )
                    if not created:
                        print(f'Student with ID {id} already exists!')                    
                    
                    # Course_registration 객체 생성 또는 가져오기
                    new_course_reg, created = Course_registration.objects.get_or_create(
                        course=specific_course,
                        course_year=specific_course.year,
                        course_semester=specific_course.semester,
                        student=new_reg_student )
                    
                    if created:
                                print(f'Student {id}, {name} registered for {specific_course.name}')
                                new_count += 1
                    else:
                        print(f'Student {id}, {name} already registered for {specific_course.name}')

                except Exception as e:
                    print(f'Error processing student with ID {id}: {e}')
                    continue

            data.append(f'{specific_course.name}: new student created: {new_count} , empty student: {empty_ids}')

        return JsonResponse(data , safe=False) 

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



def course_reg_look(request):
    try:
        access_token = get_kuopenapi_access_token()
        courses = Course.objects.all()
        data = []
        for specific_course in courses:
            course_id, course_class = specific_course.course_id.split('-')
            
            #API 요청
            api_url = "https://kuapi.korea.ac.kr/svc/course/course-registration/by-class"  # 실제 API 엔드포인트로 변경
            headers = {
                'AUTH_KEY': access_token
            }
            
            # 요청 파라미터 설정
            params = {
                'client_id': settings.KOREAUNIV_OPENAPI_CLIENT_ID,
                "year":str(specific_course.year),
                'term':str(specific_course.semester)+'R',
                'course_code': course_id,
                'course_class': course_class
            }
            
            # API 호출
            response = requests.get(api_url, headers=headers, params=params)
            # JSON 응답을 파싱
            
            response_data = response.json()

            data.append(response_data)

        return JsonResponse(data , safe=False) 

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    

def course_read_min_max_avg(request):
    try:
        # 1) 과목별 학생 수 집계
        course_stats = Course_registration.objects.select_related(
            'course', 'student'
        ).values(
            'course__course_id',
            'course__year',
            'course__semester',
            'course__name',
            'course__prof'
        ).annotate(
            student_count=Count('student', distinct=True)
        )
        
        merged_stats = []
        
        for stat in course_stats:
            # 2) 해당 과목의 학생들 github_id 수집
            student_github_ids = list(
                Course_registration.objects.filter(
                    course__course_id=stat['course__course_id'],
                    course__year=stat['course__year'],
                    course__semester=stat['course__semester']
                ).select_related('student').values_list(
                    'student__github_id', flat=True
                )
            )
            
            # 3) Repository 통계 가져오기 (Python에서 처리)
            commits = list(
                Repository.objects.filter(
                    owner_github_id__in=student_github_ids
                ).values_list('contributed_commit_count', flat=True)
            )
            
            prs = list(
                Repository.objects.filter(
                    owner_github_id__in=student_github_ids
                ).values_list('contributed_open_pr_count', 'contributed_closed_pr_count')
            )
            
            issues = list(
                Repository.objects.filter(
                    owner_github_id__in=student_github_ids
                ).values_list('contributed_open_issue_count', 'contributed_closed_issue_count')
            )
            
            stars = list(
                Repository.objects.filter(
                    owner_github_id__in=student_github_ids
                ).values_list('star_count', flat=True)
            )
            
            # 4) 통계 계산
            if commits:
                commits_arr = np.array([c or 0 for c in commits])
                commit_q1 = float(np.percentile(commits_arr, 25))
                commit_q2 = float(np.percentile(commits_arr, 50))
                commit_q3 = float(np.percentile(commits_arr, 75))
                commit_std = float(np.std(commits_arr))
                commit_min = float(np.min(commits_arr))
                commit_max = float(np.max(commits_arr))
                commit_avg = float(np.mean(commits_arr))
            else:
                commit_q1 = commit_q2 = commit_q3 = commit_std = 0
                commit_min = commit_max = commit_avg = 0
            
            if prs:
                total_prs = [(open or 0) + (closed or 0) for open, closed in prs]
                pr_arr = np.array(total_prs)
                pr_q1 = float(np.percentile(pr_arr, 25))
                pr_q2 = float(np.percentile(pr_arr, 50))
                pr_q3 = float(np.percentile(pr_arr, 75))
                pr_std = float(np.std(pr_arr))
                pr_min = float(np.min(pr_arr))
                pr_max = float(np.max(pr_arr))
                pr_avg = float(np.mean(pr_arr))
            else:
                pr_q1 = pr_q2 = pr_q3 = pr_std = 0
                pr_min = pr_max = pr_avg = 0
            
            if issues:
                total_issues = [(open or 0) + (closed or 0) for open, closed in issues]
                issue_arr = np.array(total_issues)
                issue_q1 = float(np.percentile(issue_arr, 25))
                issue_q2 = float(np.percentile(issue_arr, 50))
                issue_q3 = float(np.percentile(issue_arr, 75))
                issue_std = float(np.std(issue_arr))
                issue_min = float(np.min(issue_arr))
                issue_max = float(np.max(issue_arr))
                issue_avg = float(np.mean(issue_arr))
            else:
                issue_q1 = issue_q2 = issue_q3 = issue_std = 0
                issue_min = issue_max = issue_avg = 0
            
            if stars:
                star_arr = np.array([s or 0 for s in stars])
                star_q1 = float(np.percentile(star_arr, 25))
                star_q2 = float(np.percentile(star_arr, 50))
                star_q3 = float(np.percentile(star_arr, 75))
                star_std = float(np.std(star_arr))
                star_min = float(np.min(star_arr))
                star_max = float(np.max(star_arr))
                star_avg = float(np.mean(star_arr))
            else:
                star_q1 = star_q2 = star_q3 = star_std = 0
                star_min = star_max = star_avg = 0
            
            merged_stats.append({
                "course_id": stat['course__course_id'],
                "year": stat['course__year'],
                "semester": stat['course__semester'],
                "course_name": stat['course__name'],
                "prof": stat['course__prof'],
                "student_count": stat['student_count'],
                "commit_q1": commit_q1,
                "commit_q2": commit_q2,
                "commit_q3": commit_q3,
                "commit_std": commit_std,
                "commit_min": commit_min,
                "commit_max": commit_max,
                "commit_avg": commit_avg,
                "pr_q1": pr_q1,
                "pr_q2": pr_q2,
                "pr_q3": pr_q3,
                "pr_std": pr_std,
                "pr_min": pr_min,
                "pr_max": pr_max,
                "pr_avg": pr_avg,
                "issue_q1": issue_q1,
                "issue_q2": issue_q2,
                "issue_q3": issue_q3,
                "issue_std": issue_std,
                "issue_min": issue_min,
                "issue_max": issue_max,
                "issue_avg": issue_avg,
                "star_q1": star_q1,
                "star_q2": star_q2,
                "star_q3": star_q3,
                "star_std": star_std,
                "star_min": star_min,
                "star_max": star_max,
                "star_avg": star_avg
            })
        
        return JsonResponse({"group_stats": merged_stats}, safe=False)
    
    except Exception as e:
        return JsonResponse({"error": str(e)}, safe=False)

def course_read_db_total_excel(request):
    try:
        data = []
        courses = Course.objects.all()

        for course in courses:

            total_commits = 0
            total_prs = 0
            total_issues = 0
            total_stars = 0

            # Gather course related repos' id
            course_related_repos_id = Course_project.objects.filter(course_id=course.id).values_list('repo_id', flat=True)
            print(f'{course.name}:[{course.course_id}] - repo_count: {len(course_related_repos_id)}')

            total_course_repos = Repository.objects.filter(id__in=course_related_repos_id)

            for course_repo in total_course_repos:
                if course_repo.forked is True:
                    continue
                else:
                    total_commits += course_repo.commit_count or 0
                    total_issues += ((course_repo.open_issue_count or 0) + (course_repo.closed_issue_count or 0))
                    total_prs += ((course_repo.open_pr_count or 0) + (course_repo.closed_pr_count or 0))
                    total_stars += course_repo.star_count or 0

            # Calculate the average commits
            student_count = Course_registration.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            ).count()

            avg_commits = round(total_commits / student_count, 2) if student_count > 0 else 0

            # Calculate the number of repositories
            repository_count = Course_project.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            ).count()

            # Calculate the number of contributors
            course_repo_ids = Course_project.objects.filter(
                course=course,
                course_year=course.year,
                course_semester=course.semester
            )

            contributor_count = 0
            for course_repo_id in course_repo_ids:
                if course_repo_id.repo.forked == False:
                    contributors_list_str = course_repo_id.repo.contributors
                    cleaned_contributors_list = ''.join(contributors_list_str.split())  # Remove all whitespace
                    if cleaned_contributors_list == '':
                        continue

                    contributors_number = cleaned_contributors_list.count(',') + 1  # Count contributors by commas
                    contributor_count += contributors_number

            # Gather all the data
            data.append({
                "course_id": course.course_id,
                "year": course.year,
                "semester": course.semester,
                "name": course.name,
                "prof": course.prof,
                "ta": course.ta,
                "student_count": course.student_count,
                "total_commits": total_commits,
                "total_issues": total_issues,
                "total_prs": total_prs,
                "total_stars": total_stars,
                "avg_commits": avg_commits,
                "repository_count": repository_count,
                "contributor_count": contributor_count
            })

        # Create Excel file
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Course Data"

        # Write headers
        headers = [
            "Course ID", "Year", "Semester", "Name", "Professor", "TA", "Student Count",
            "Total Commits", "Total Issues", "Total PRs", "Total Stars", "Average Commits",
            "Repository Count", "Contributor Count"
        ]
        sheet.append(headers)

        # Write data rows
        for row in data:
            sheet.append([
                row["course_id"], row["year"], row["semester"], row["name"], row["prof"], row["ta"],
                row["student_count"], row["total_commits"], row["total_issues"], row["total_prs"],
                row["total_stars"], row["avg_commits"], row["repository_count"], row["contributor_count"]
            ])

        # Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:  # Necessary to avoid issues with NoneType
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Save workbook to BytesIO
        workbook.save(output)
        output.seek(0)

        # Return as Excel file response
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="course_data.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)

def course_student_db_excel(request):
    try:
        # 데이터를 가져오는 함수 호출
        response = student_read_course_info(request)

        # JsonResponse에서 데이터 추출
        if hasattr(response, 'content'):
            student_course_list = json.loads(response.content.decode('utf-8'))
        else:
            raise ValueError("Invalid response format from student_read_course_info")

        # 반환 데이터 형식 확인
        if not isinstance(student_course_list, list):
            raise ValueError("Invalid data format: Expected a list of dictionaries.")

    except Exception as e:
        return HttpResponse(f"Error processing student data: {str(e)}", status=500)

    # course_id, year, semester로 데이터 그룹화
    grouped_data = {}
    for student in student_course_list:
        try:
            course_id = student.get("course_id", "").strip() or "기타"
            year = student.get("year") or 0  # 기본값을 0으로 설정
            semester = student.get("semester") or 0  # 기본값을 0으로 설정

            # Ensure year and semester are integers
            year = int(year) if year else 0
            semester = int(semester) if semester else 0

            group_key = f"{course_id}_{year}_{semester}"
            if group_key not in grouped_data:
                grouped_data[group_key] = []
            
            # Repository Name 가져오기
            try:
                student_id = student.get("id")
                specific_student = Student.objects.get(id=student_id)
                course_info= Course.objects.get(course_id=course_id, year=year, semester=semester)
                course_repositories = Course_project.objects.filter(course = course_info)
                for course_repo in course_repositories:
                    if course_repo.repo.owner_github_id == specific_student.github_id :
                        repository_name = course_repo.repo.name
                        break
                    else :
                        continue
            except ObjectDoesNotExist:
                repository_name = "Unknown"

            student["repository_name"] = repository_name
            grouped_data[group_key].append(student)
        except AttributeError as e:
            return HttpResponse(f"Error processing student record: {str(e)}", status=500)

    # 엑셀 파일 생성
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # 기본 시트 제거

    for group_key, students in grouped_data.items():
        # 시트 이름 생성 (길이 제한 고려)
        sheet_name = group_key[:31]  # 시트 이름은 최대 31자
        sheet = workbook.create_sheet(title=sheet_name)

        # 헤더 작성
        headers = [
            "ID", "GitHub ID", "Name", "Department", "Enrollment",
            "Year", "Semester", "Course Name", "Commit", "PR", "Issue",
            "Num Repos", "Star Count", "Professor", "Total Contributors",  "Repository Names"
        ]
        sheet.append(headers)
        for col in sheet[1]:
            col.font = Font(bold=True)

        # 학생 데이터 삽입
        for student in students:
            row = [
                student.get("id"),
                student.get("github_id"),
                student.get("name"),
                student.get("department"),
                student.get("enrollment"),
                student.get("year"),
                student.get("semester"),
                student.get("course_name"),
                student.get("commit"),
                student.get("pr"),
                student.get("issue"),
                student.get("num_repos"),
                student.get("star_count"),
                student.get("prof"),
                student.get("total_contributors"),
                student.get("repository_name")
            ]
            sheet.append(row)

    # 엑셀 파일 저장
    workbook.save(output)
    output.seek(0)

    # HTTP 응답으로 엑셀 파일 반환
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename=\"course_students.xlsx\"'
    return response
# def course_student_db_excel(request):
#     try:
#         data = []
#         courses = Course.objects.all()

#         for course in courses:
            
#             course_id = course.course_id
#             year = course.year
#             semester = course.semester


#             # Gather course related repos' id
#             course_related_repos_id = Course_project.objects.filter(course_id=course.id).values_list('repo_id', flat=True)
#             print(f'{course.name}:[{course.course_id}] - repo_count: {len(course_related_repos_id)}')
            
            
#             total_course_repos = Repository.objects.filter(id__in=course_related_repos_id)

#             for course_repo in total_course_repos :

#                 if course_repo.forked is True :
#                     continue            
#                 else :
                    
#                     student = Student.objects.get(github_id = course_repo.owner_github_id)                
#                     student_id = student.id 
#                     student_github_id = student.github_id
#                     student_name = student.name 
#                     student_department = student.department
#                     student_enrollment = student.enrollment
                    
#                     commit_couunt = course_repo.commit_count or 0 
#                     issue_count =  ((course_repo.open_issue_count or 0) + (course_repo.closed_issue_count or 0))
#                     pr_count = ((course_repo.open_pr_count or 0) + (course_repo.closed_pr_count or 0))
#                     star_count = course_repo.star_count or 0 
            

#                     if course_repo.forked == False :
#                         contributors_list_str = course_repo.contributors
#                         cleaned_contributors_list = ''.join(contributors_list_str.split())  # 모든 공백 제거
#                         if cleaned_contributors_list == '':
#                             continue
#                         contributors_number = cleaned_contributors_list.count(',') + 1 # , 코마갯수로 contributor 세기

#     except Exception as e:
#         return JsonResponse({"status": "Error", "message": str(e)}, status=500)


def only_for_course_student_db_excel(request): #과목 관점에서 읽기 
    try:
        data = []
        students = Student.objects.all()
        for student in students:
            try:
                # Proceed for one specific student             
                total_commit = 0 
                total_pr = 0
                total_issue = 0
                total_repo = 0
                total_star = 0

                etc_total_commit = 0 
                etc_total_pr = 0
                etc_total_issue = 0
                etc_total_repo = 0
                etc_total_star = 0
                etc_total_contributor=0

                # Get all repositories for a specific student
                student_repos = Repository.objects.filter(owner_github_id=student.github_id)
                
                # 특정 학생의 수강 목록들 가져옴
                course_reg_list = Course_registration.objects.filter(student=student)

                # 특정 학생의 과목 관련된 레포지토리들을 가져옴.
                courses_repos = Course_project.objects.filter(repo__in=student_repos)
                
                # 특정 학생이 듣는 모든 course_id 테이블 및 딕셔너리 생성
                course_ids = []
                for course_reg in course_reg_list:
                    course_ids.append(course_reg.course.course_id)

                # 딕셔너리 초기화
                course_dict = {course_id: {'commit': 0, 'pr': 0, 'issue': 0, 'repo': 0, 'star': 0, 'contributors': 0,} for course_id in course_ids}

                courses_project_repos = []

                for c in courses_repos:
                    courses_project_repos.append(c.repo)

                # 각 repo에 대한 정보 추가
                for repo in student_repos:
                    sanitized_url = repo.url.replace(":", "")

                    if repo in courses_project_repos:  # 특정 학생의 현재 가리키는 repo가 과목과 관련이 있는 경우
                        specific_course = Course_project.objects.get(repo=repo)
                        course_id = specific_course.course.course_id

                        # Commit 수 합산
                        total_commit_count = Repository.objects.filter(
                            id = specific_course.repo.id,
                            owner_github_id=student.github_id
                        ).aggregate(total_commits=Sum('contributed_commit_count'))['total_commits'] or 0
                        
                        # 합산한 값을 course_dict에 추가
                        course_dict[course_id]['commit'] += total_commit_count
                        
                        # PR 수 합산
                        course_dict[course_id]['pr'] += ( repo.contributed_open_pr_count or 0) +  (repo.contributed_closed_pr_count or 0)
                        
                        # Issue 수 합산
                        course_dict[course_id]['issue'] += (repo.contributed_open_issue_count or 0) + (repo.contributed_closed_issue_count or 0)
                        
                        # Repo 수 합산
                        course_dict[course_id]['repo'] += 1 
                        
                        # Star 수 합산
                        course_dict[course_id]['star'] += Repository.objects.get(id=repo.id).star_count or 0
                        print(repo.url)

                        course_dict[course_id]['contributors'] += Repo_contributor.objects.filter(repo_url = sanitized_url).count() or 0 

                    else:  # 특정 학생의 repo가 과목과 관련 없는 경우
                        etc_total_commit += repo.contributed_commit_count or 0 
                        etc_total_pr +=  ( repo.contributed_open_pr_count or 0) +  (repo.contributed_closed_pr_count or 0)
                        etc_total_issue += (repo.contributed_open_issue_count or 0) + (repo.contributed_closed_issue_count or 0)
                        etc_total_repo += 1
                        etc_total_star += Repository.objects.get(id=repo.id).star_count or 0
                        etc_total_contributor += Repo_contributor.objects.filter(repo_url = sanitized_url).count() or 0

                # 각 과목에 대한 정보 추가
                for course_id, course_count in course_dict.items():
                    course = Course.objects.get(course_id=course_id)

                    course_info = {
                        "id": student.id,
                        "github_id": student.github_id,
                        "name": student.name,
                        "department": student.department,
                        "enrollment": student.enrollment,
                        "year": course.year,
                        "semester": course.semester,
                        "course_name": course.name,
                        "commit": course_count['commit'],
                        "pr": course_count['pr'],
                        "issue": course_count['issue'],
                        "num_repos": course_count['repo'],
                        "star_count": course_count['star'],
                        "prof": course.prof,
                        "course_id": course.course_id,
                        "total_contributors": course_count['contributors']
                    }
                    total_commit += course_count['commit']
                    total_pr += course_count['pr']
                    total_issue += course_count['issue']
                    total_repo += course_count['repo']
                    total_star += course_count['star']

                    data.append(course_info)

                # 기타 정보 추가
                etc_info = {
                    "id": student.id,
                    "github_id": student.github_id,
                    "name": student.name,
                    "department": student.department,
                    "enrollment": student.enrollment,
                    "year": "",
                    "semester": "",
                    "course_name": "기타",
                    "commit": etc_total_commit,
                    "pr": etc_total_pr,
                    "issue": etc_total_issue,
                    "num_repos": etc_total_repo,
                    "star_count": etc_total_star,
                    "prof": "",
                    "course_id": "",
                    "total_contributors": etc_total_contributor
                }

                total_commit += etc_total_commit
                total_pr += etc_total_pr
                total_issue += etc_total_issue
                total_repo += etc_total_repo
                total_star += etc_total_star

                data.append(etc_info)               

            except Exception as e:
                print(f'Error processing student {student.name}: {e}')
                continue       

        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"status": "Error", "message": str(e)}, status=500)